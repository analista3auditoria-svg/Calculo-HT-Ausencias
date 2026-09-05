import os
import io
import datetime
import warnings
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st

# Ocultar advertencias no críticas
warnings.filterwarnings('ignore', category=UserWarning)

# Configuración de página en Streamlit
st.set_page_config(
    page_title="Auditor TS - Casalimpia", 
    page_icon="🧼", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── ESTILOS Y PALETA DE COLORES CORPORATIVOS CASALIMPIA ──────────────────
st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');
        
        * {
            font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, sans-serif;
        }

        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        
        [data-testid="stSidebar"] {
            display: block !important;
            visibility: visible !important;
        }
        [data-testid="stSidebarNav"] {
            display: block !important;
        }
        
        .main {
            background-color: #f8fafc;
        }

        .header-brand {
            background: linear-gradient(135deg, #ffffff 0%, #f0f7ff 100%);
            border-left: 6px solid #00529B;
            border-top: 1px solid #e2e8f0;
            border-right: 1px solid #e2e8f0;
            border-bottom: 1px solid #e2e8f0;
            border-radius: 16px;
            padding: 24px 32px;
            margin-bottom: 24px;
            box-shadow: 0 4px 20px -2px rgba(0, 82, 155, 0.06);
        }
        .header-brand-content {
            display: flex;
            align-items: center;
            gap: 24px;
        }
        .header-brand-content img {
            height: 58px;
            width: auto;
            object-fit: contain;
        }
        .title-text h1 {
            color: #00529B !important;
            font-size: 24px !important;
            font-weight: 700 !important;
            margin: 0 !important;
            padding: 0 !important;
            line-height: 1.2 !important;
            letter-spacing: -0.02em;
        }
        .title-text p {
            color: #64748b !important;
            font-size: 14px !important;
            margin: 4px 0 0 0 !important;
            padding: 0 !important;
            font-weight: 500;
        }

        .card-container {
            background-color: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 14px;
            padding: 20px 24px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px -2px rgba(0,0,0,0.04);
        }
        .section-header {
            display: flex;
            align-items: center;
            justify-content: space-between;
            color: #00529B;
            font-size: 16px;
            font-weight: 700;
            margin-bottom: 14px;
            padding-bottom: 8px;
            border-bottom: 2px solid #f0f7ff;
        }

        .file-status-ok {
            background-color: #dcfce7;
            color: #15803d;
            font-size: 12px;
            font-weight: 600;
            padding: 4px 10px;
            border-radius: 20px;
        }
        .file-status-pending {
            background-color: #f1f5f9;
            color: #64748b;
            font-size: 12px;
            font-weight: 500;
            padding: 4px 10px;
            border-radius: 20px;
        }

        .kpi-card {
            background-color: #ffffff;
            border-radius: 16px;
            padding: 20px 24px;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.03);
            border: 1px solid #e2e8f0;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }
        .kpi-card:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 16px rgba(0, 0, 0, 0.06);
        }
        .kpi-card-danger {
            border-left: 6px solid #dc2626;
            background: linear-gradient(135deg, #ffffff 0%, #fef2f2 100%);
        }
        .kpi-card-warning {
            border-left: 6px solid #d97706;
            background: linear-gradient(135deg, #ffffff 0%, #fffbeb 100%);
        }
        .kpi-card-info {
            border-left: 6px solid #00529B;
            background: linear-gradient(135deg, #ffffff 0%, #f0f7ff 100%);
        }
        .kpi-title {
            font-size: 13px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin-bottom: 8px;
        }
        .kpi-title-danger { color: #991b1b; }
        .kpi-title-warning { color: #92400e; }
        .kpi-title-info { color: #00529B; }

        .kpi-value {
            font-size: 36px;
            font-weight: 800;
            line-height: 1;
            margin-bottom: 6px;
        }
        .kpi-value-danger { color: #dc2626; }
        .kpi-value-warning { color: #d97706; }
        .kpi-value-info { color: #00529B; }

        .kpi-subtitle {
            font-size: 12px;
            font-weight: 500;
            color: #64748b;
        }

        [data-testid="stFileUploader"] {
            padding: 0px;
            margin-bottom: 12px;
        }
        [data-testid="stFileUploaderDropzone"] {
            padding: 10px 14px !important;
            border: 1.5px dashed #cbd5e1 !important;
            border-radius: 10px !important;
            background-color: #f8fafc !important;
            display: flex !important;
            flex-direction: row !important;
            align-items: center !important;
            justify-content: flex-start !important;
            gap: 12px !important;
            min-height: 46px !important;
            transition: all 0.2s ease;
        }
        [data-testid="stFileUploaderDropzone"]:hover {
            background-color: #f0f7ff !important;
            border-color: #00529B !important;
        }
        [data-testid="stFileUploaderDropzone"] section,
        [data-testid="stFileUploaderDropzone"] small,
        [data-testid="stFileUploaderDropzone"] svg {
            display: none !important;
        }
        [data-testid="stFileUploaderDropzone"]::before {
            content: "📄 Cargar" !important;
            display: inline-block !important;
            background-color: #00529B !important;
            color: #ffffff !important;
            border: none !important;
            border-radius: 6px !important;
            padding: 5px 12px !important;
            font-size: 12.5px !important;
            font-weight: 600 !important;
            cursor: pointer !important;
        }
        [data-testid="stFileUploaderDropzone"] button {
            display: none !important;
        }

        div.stButton > button:first-child {
            background: linear-gradient(135deg, #00529B 0%, #003366 100%) !important;
            color: white !important;
            border-radius: 10px !important;
            padding: 12px 28px !important;
            font-weight: 700 !important;
            font-size: 16px !important;
            border: none !important;
            box-shadow: 0 4px 12px rgba(0, 82, 155, 0.25) !important;
            transition: all 0.2s ease !important;
            width: 100%;
        }
        div.stButton > button:first-child:hover {
            background: linear-gradient(135deg, #003366 0%, #002244 100%) !important;
            box-shadow: 0 6px 16px rgba(0, 82, 155, 0.35) !important;
            transform: translateY(-1px);
        }

        div.stDownloadButton > button:first-child {
            background: linear-gradient(135deg, #059669 0%, #047857 100%) !important;
            color: white !important;
            border-radius: 10px !important;
            padding: 12px 28px !important;
            font-weight: 700 !important;
            font-size: 16px !important;
            border: none !important;
            box-shadow: 0 4px 12px rgba(5, 150, 105, 0.25) !important;
            width: 100%;
        }
        div.stDownloadButton > button:first-child:hover {
            background: linear-gradient(135deg, #047857 0%, #065f46 100%) !important;
            box-shadow: 0 6px 16px rgba(5, 150, 105, 0.35) !important;
        }
    </style>

    <div class="header-brand">
        <div class="header-brand-content">
            <img src="https://cdn1.totalcommerce.cloud/casalimpia/web_content/assets/logo-casa-limpia.svg" alt="Casalimpia Logo" />
            <div class="title-text">
                <h1>Auditor TS & Módulo GeoVictoria</h1>
                <p>Plataforma Corporativa de Procesamiento y Auditoría de Tiempos y Ausentismos</p>
            </div>
        </div>
    </div>
""", unsafe_allow_html=True)


# ─── FUNCIONES DE APOYO Y PROCESAMIENTO ───────────────────────────────────

def obtener_val_iloc(row, index_col):
    if len(row) > index_col and pd.notna(row.iloc[index_col]):
        return str(row.iloc[index_col]).strip()
    return ""

def convertir_a_hora(val):
    if not val:
        return None
    try:
        dt = pd.to_datetime(val, format='%H:%M', errors='coerce')
        if pd.isna(dt):
            dt = pd.to_datetime(val, errors='coerce')
        if pd.notna(dt):
            return dt.time()
    except Exception:
        pass
    return None

def procesar_plantilla_geovictoria(
    file_entrada, sheet_entrada, sheet_festivos,
    file_operativa, sheet_operativa,
    file_novasoft, sheet_novasoft,
    file_sic, sheet_sic,
    file_maestro, sheet_maestro,
    file_historial, sheet_historial,
    file_supernumerario, sheet_supernumerario,
    contrato_principal,
    fecha_ini_sup, fecha_fin_sup
):
    df_marc = pd.read_excel(file_entrada, sheet_name=sheet_entrada)
    
    operativa_dict = {}
    if file_operativa:
        try:
            excel_op = pd.ExcelFile(file_operativa)
            target_sheet = sheet_operativa
            for name in excel_op.sheet_names:
                if name.strip().upper() == sheet_operativa.strip().upper():
                    target_sheet = name
                    break
            df_op = pd.read_excel(file_operativa, sheet_name=target_sheet)
            
            if not df_op.empty:
                df_op['Cédula_Str'] = df_op.apply(lambda r: obtener_val_iloc(r, 0).replace('.0', ''), axis=1)
                df_op['Fecha_Dt'] = pd.to_datetime(df_op.iloc[:, 2], dayfirst=True, errors='coerce')
                df_op['Val_C'] = df_op.apply(lambda r: obtener_val_iloc(r, 8), axis=1)
                
                for _, row_op in df_op.iterrows():
                    ced_op = row_op['Cédula_Str']
                    f_op = row_op['Fecha_Dt']
                    val_c = row_op['Val_C']
                    if ced_op and pd.notna(f_op):
                        operativa_dict[(ced_op, f_op.date())] = val_c
        except Exception as e:
            st.warning(f"⚠️ No se pudo procesar la hoja '{sheet_operativa}' de la Base Operativa: {e}")

    df_super_filtrado = pd.DataFrame()
    if file_supernumerario:
        try:
            excel_sup = pd.ExcelFile(file_supernumerario)
            target_sheet_sup = sheet_supernumerario
            for name in excel_sup.sheet_names:
                if name.strip().upper() == sheet_supernumerario.strip().upper():
                    target_sheet_sup = name
                    break
            df_sup_raw = pd.read_excel(file_supernumerario, sheet_name=target_sheet_sup)
            
            if not df_sup_raw.empty and df_sup_raw.shape[1] > 12:
                col_m_val = df_sup_raw.iloc[:, 12].astype(str).str.strip().str.upper()
                contrato_target = str(contrato_principal).strip().upper()
                mask = col_m_val.str.contains(contrato_target, regex=False, na=False)

                if fecha_ini_sup and fecha_fin_sup:
                    fechas_col_b = pd.to_datetime(df_sup_raw.iloc[:, 1], dayfirst=True, errors='coerce').dt.date
                    mask = mask & (fechas_col_b >= fecha_ini_sup) & (fechas_col_b <= fecha_fin_sup)

                df_super_filtrado = df_sup_raw[mask].copy()
        except Exception as e:
            st.warning(f"⚠️ No se pudo procesar la BD Supernumerario (hoja '{sheet_supernumerario}'): {e}")

    df_nova = pd.read_excel(file_novasoft, sheet_name=sheet_novasoft) if file_novasoft else pd.DataFrame()
    df_sic = pd.read_excel(file_sic, sheet_name=sheet_sic) if file_sic else pd.DataFrame()
    df_maestro = pd.read_excel(file_maestro, sheet_name=sheet_maestro) if file_maestro else pd.DataFrame()
    df_hist = pd.read_excel(file_historial, sheet_name=sheet_historial) if file_historial else pd.DataFrame()

    set_festivos = set()
    if file_entrada:
        try:
            df_festivos = pd.read_excel(file_entrada, sheet_name=sheet_festivos)
            if not df_festivos.empty:
                fechas_fest = pd.to_datetime(df_festivos.iloc[:, 0], dayfirst=True, errors='coerce').dropna()
                set_festivos = set(fechas_fest.dt.date)
        except Exception as e:
            st.warning(f"⚠️ Nota: No se pudo cargar la hoja '{sheet_festivos}' ({e}). Se continuará sin marcar festivos.")

    df_marc['Cédula_Str'] = df_marc.apply(lambda r: obtener_val_iloc(r, 2).replace('.0', ''), axis=1)
    
    if not df_hist.empty:
        df_hist['Cédula_Str'] = df_hist.apply(lambda r: obtener_val_iloc(r, 0).replace('.0', ''), axis=1)
        df_hist['Centro_Costo'] = df_hist.apply(lambda r: obtener_val_iloc(r, 2), axis=1)
        df_hist['Fecha_Inicio'] = pd.to_datetime(df_hist.iloc[:, 3], dayfirst=True, errors='coerce') if df_hist.shape[1] > 3 else pd.NaT
        if df_hist.shape[1] > 4:
            df_hist['Fecha_Fin_Raw'] = pd.to_datetime(df_hist.iloc[:, 4], dayfirst=True, errors='coerce')
            fecha_dummy = pd.to_datetime('2001-01-01')
            df_hist['Fecha_Fin'] = df_hist['Fecha_Fin_Raw'].apply(
                lambda x: pd.to_datetime('2099-12-31') if (pd.isna(x) or x == fecha_dummy) else x
            )
        else:
            df_hist['Fecha_Fin'] = pd.to_datetime('2099-12-31')
        df_hist['Frente_Trabajo'] = df_hist.apply(lambda r: obtener_val_iloc(r, 5), axis=1) if df_hist.shape[1] > 5 else df_hist['Centro_Costo']

    if not df_nova.empty:
        df_nova['Cédula_Str'] = df_nova.apply(lambda r: obtener_val_iloc(r, 0).replace('.0', ''), axis=1)
        df_nova['Concepto'] = df_nova.apply(lambda r: obtener_val_iloc(r, 2), axis=1)
        df_nova['Fecha_Inicio'] = pd.to_datetime(df_nova.iloc[:, 3], dayfirst=True, errors='coerce') if df_nova.shape[1] > 3 else pd.NaT
        df_nova['Fecha_Fin'] = pd.to_datetime(df_nova.iloc[:, 4], dayfirst=True, errors='coerce') if df_nova.shape[1] > 4 else pd.NaT
        df_nova['Codigo_Novasoft'] = df_nova.apply(lambda r: obtener_val_iloc(r, 9), axis=1)

    if not df_sic.empty:
        df_sic['Cédula_Str'] = df_sic.apply(lambda r: obtener_val_iloc(r, 9).replace('.0', ''), axis=1)
        df_sic['Proceso'] = df_sic.apply(lambda r: obtener_val_iloc(r, 1), axis=1)
        df_sic['Estado'] = df_sic.apply(lambda r: obtener_val_iloc(r, 32), axis=1)
        df_sic = df_sic[df_sic['Estado'].str.lower() == 'nomina'].copy()
        if df_sic.shape[1] > 4:
            df_sic['Fecha_Inicio'] = pd.to_datetime(df_sic.iloc[:, 4], dayfirst=True, errors='coerce')
        if df_sic.shape[1] > 5:
            df_sic['Fecha_Fin'] = pd.to_datetime(df_sic.iloc[:, 5], dayfirst=True, errors='coerce')

    maestro_dict = {}
    if not df_maestro.empty:
        df_maestro['Cédula_Str'] = df_maestro.apply(lambda r: obtener_val_iloc(r, 1).replace('.0', ''), axis=1)
        df_maestro['F_INGRESO'] = pd.to_datetime(df_maestro.iloc[:, 26], dayfirst=True, errors='coerce') if df_maestro.shape[1] > 26 else pd.NaT
        df_maestro['F_RETIRO'] = pd.to_datetime(df_maestro.iloc[:, 27], dayfirst=True, errors='coerce') if df_maestro.shape[1] > 27 else pd.NaT
        for _, row_m in df_maestro.iterrows():
            if row_m['Cédula_Str']:
                maestro_dict[row_m['Cédula_Str']] = (row_m['F_INGRESO'], row_m['F_RETIRO'])

    file_entrada.seek(0)
    wb = openpyxl.load_workbook(file_entrada, data_only=False)
    ws = wb[sheet_entrada]
    ws.views.sheetView[0].showGridLines = True

    encabezados_estilos = [
        ("AY1", "Fecha Ori", "D0E1F9", "002244", True),
        ("AZ1", "Dia", "D0E1F9", "002244", True),
        ("BA1", "Entrada2", "FCE8E6", "A80000", True),
        ("BB1", "Salida2", "FCE8E6", "A80000", True),
        ("BC1", "Cant HT", "E2E8F0", "1E293B", True),
        ("BD1", "HT", "E2E8F0", "1E293B", True),
        ("BE1", "Compensatorio", "00529B", "FFFFFF", True),
        ("BF1", "Ausencias / Marcaciones Erroneas", "990000", "FFFFFF", True),
        ("BG1", "Recargo Dominical No Compensado", "DCFCE7", "14532D", True),
        ("BH1", "Recargo Dominical Compensado", "FEF08A", "713F12", True),
        ("BI1", "Recargo Festivo", "FFEDD5", "7C2D12", True),
        ("BJ1", "Recargo Nocturno 0.35%", "BAE6FD", "0369A1", True),
        ("BK1", "Horas Extras Diurnas 1.25%", "E9D5FF", "581C87", True),
        ("BL1", "Hora Extra Diurna Dom/Fest", "BBF7D0", "166534", True),
        ("BM1", "Horas Extras Nocturnas 1.75%", "FECDD3", "9F1239", True),
        ("BN1", "Hora Extra Dominical o Festiva Nocturna", "C6EFCE", "064E3B", True),
        ("BO1", "CCCO", "003366", "FFFFFF", True),
        ("BP1", "Frente de trabajo", "003366", "FFFFFF", True),
        ("BQ1", "Centro de costos", "003366", "FFFFFF", True),
        ("BR1", "Fecha Ingreso", "00529B", "FFFFFF", True),
        ("BS1", "Validar Fingreso", "00529B", "FFFFFF", True),
        ("BT1", "F Retiro", "00529B", "FFFFFF", True),
        ("BU1", "Validar F Retiro", "00529B", "FFFFFF", True),
        ("BV1", "Ausentismo Novasoft", "990000", "FFFFFF", True),
        ("BW1", "Codigo novasoft", "00529B", "FFFFFF", True),
        ("BX1", "Ausentismo Sic", "00529B", "FFFFFF", True),
        ("BY1", "Ausentismo", "C00000", "FFFFFF", True),
        ("BZ1", "Compensado", "00529B", "FFFFFF", True),
        ("CA1", "Ausentismo Real", "1E4620", "FFFFFF", True)
    ]

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    fill_ausencia_rojo = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    font_ausencia_blanco = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    for celda_ref, titulo, color_bg, color_fg, es_negrita in encabezados_estilos:
        celda = ws[celda_ref]
        celda.value = titulo
        celda.fill = PatternFill(start_color=color_bg, end_color=color_bg, fill_type="solid")
        celda.font = Font(name="Calibri", size=10, bold=es_negrita, color=color_fg)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = thin_border

    dias_semana_es = {0: "lunes", 1: "martes", 2: "miércoles", 3: "jueves", 4: "viernes", 5: "sábado", 6: "domingo"}
    hist_dict = {ced: grp for ced, grp in df_hist.groupby('Cédula_Str')} if not df_hist.empty else {}
    nova_dict = {ced: grp for ced, grp in df_nova.groupby('Cédula_Str')} if not df_nova.empty else {}
    sic_dict = {ced: grp for ced, grp in df_sic.groupby('Cédula_Str')} if not df_sic.empty else {}
    fecha_minima_valida = pd.to_datetime('1900-01-01')

    progress_bar = st.progress(0)
    status_text = st.empty()
    total_filas = len(df_marc)

    conteo_ausencias = 0
    conteo_p = 0
    registros_ausencias = []
    marcaciones_ht_dict = {}

    hora_corte_nocturna = datetime.time(20, 0)

    for idx, row in df_marc.iterrows():
        i = idx + 2

        val_e = obtener_val_iloc(row, 4)
        fecha_ori = None
        if len(val_e) >= 10:
            try:
                fecha_ori = pd.to_datetime(val_e[-10:], dayfirst=True)
            except Exception:
                fecha_ori = None
        
        celda_ay = ws[f'AY{i}']
        if fecha_ori is not pd.NaT and fecha_ori is not None:
            celda_ay.value = fecha_ori.date()
            celda_ay.number_format = 'DD/MM/YYYY'
        else:
            celda_ay.value = val_e

        dia_nombre = ""
        if fecha_ori and pd.notna(fecha_ori):
            dia_nombre = dias_semana_es[fecha_ori.weekday()]
            if fecha_ori.date() in set_festivos:
                dia_nombre += " Festivo"
        
        ws[f'AZ{i}'].value = dia_nombre

        h_val, j_val = obtener_val_iloc(row, 7), obtener_val_iloc(row, 9)
        k_val, m_val = obtener_val_iloc(row, 10), obtener_val_iloc(row, 12)

        hora_h = convertir_a_hora(h_val)
        hora_j = convertir_a_hora(j_val)
        hora_k = convertir_a_hora(k_val)
        hora_m = convertir_a_hora(m_val)

        celda_ba = ws[f'BA{i}']
        celda_bb = ws[f'BB{i}']

        if hora_k is not None and hora_k > hora_corte_nocturna:
            celda_ba.value = hora_k
            celda_ba.number_format = 'hh:mm:ss AM/PM'

            if hora_j is not None:
                celda_bb.value = hora_j
                celda_bb.number_format = 'hh:mm:ss AM/PM'
            else:
                celda_bb.value = ""
        else:
            horas_validas = [dt for dt in [hora_h, hora_j, hora_k, hora_m] if dt is not None]

            if len(horas_validas) > 0:
                celda_ba.value = min(horas_validas)
                celda_ba.number_format = 'hh:mm:ss AM/PM'
            else:
                celda_ba.value = ""

            if len(horas_validas) > 1:
                celda_bb.value = max(horas_validas)
                celda_bb.number_format = 'hh:mm:ss AM/PM'
            else:
                celda_bb.value = ""

        val_ba = celda_ba.value
        val_bb = celda_bb.value

        ws[f'BC{i}'].value = f'=IF(OR(BA{i}="",BB{i}=""),"",MOD(BB{i}-BA{i},1))'
        ws[f'BC{i}'].number_format = '[h]:mm'
        ws[f'BD{i}'] = f'=IFERROR(ROUND(BC{i}*24,1),"")'

        f_val = obtener_val_iloc(row, 5)
        ws[f'BE{i}'] = "C" if f_val == "Descanso compensatorio" else ""

        has_ba = val_ba is not None and val_ba != ""
        has_bb = val_bb is not None and val_bb != ""

        if has_ba and has_bb:
            val_bf = ""
        elif not has_ba and not has_bb:
            if dia_nombre.lower() == "domingo" or "festivo" in dia_nombre.lower():
                val_bf = "Descanso"
            else:
                val_bf = "Ausencia"
        else:
            val_bf = "P"

        ws[f'BF{i}'] = val_bf

        ws[f'BG{i}'] = f'=AM{i}+AO{i}'
        ws[f'BH{i}'] = f'=AI{i}+AK{i}'
        ws[f'BI{i}'] = f'=AQ{i}+AS{i}+AU{i}+AW{i}'
        ws[f'BJ{i}'] = f'=AG{i}+AK{i}+AS{i}+AO{i}+AW{i}'
        ws[f'BK{i}'] = f'=U{i}'
        ws[f'BL{i}'] = f'=Y{i}+AC{i}'
        ws[f'BM{i}'] = f'=W{i}'
        ws[f'BN{i}'] = f'=AA{i}+AE{i}'

        cedula_emp = row['Cédula_Str']
        val_bo = ""
        if cedula_emp in hist_dict and fecha_ori and pd.notna(fecha_ori):
            sub_hist = hist_dict[cedula_emp]
            match_ccco = sub_hist[
                (sub_hist['Fecha_Inicio'] <= fecha_ori) &
                (sub_hist['Fecha_Fin'] >= fecha_ori) &
                (sub_hist['Centro_Costo'] != contrato_principal)
            ]
            if not match_ccco.empty:
                val_bo = "CCCO"

        ws[f'BO{i}'] = val_bo

        val_bp_cc, val_bq_cc = "", ""
        if cedula_emp in hist_dict and fecha_ori and pd.notna(fecha_ori):
            sub_hist = hist_dict[cedula_emp]
            match_hist = sub_hist[
                (sub_hist['Fecha_Inicio'] <= fecha_ori) &
                (sub_hist['Fecha_Fin'] >= fecha_ori)
            ]
            if not match_hist.empty:
                val_bp_cc = match_hist.iloc[0]['Frente_Trabajo']
                val_bq_cc = match_hist.iloc[0]['Centro_Costo']
            else:
                ult_registro = sub_hist.sort_values(by='Fecha_Inicio', ascending=False)
                if not ult_registro.empty and fecha_ori > ult_registro.iloc[0]['Fecha_Fin']:
                    val_bp_cc = ult_registro.iloc[0]['Frente_Trabajo']
                    val_bq_cc = ult_registro.iloc[0]['Centro_Costo']

        ws[f'BP{i}'] = val_bp_cc
        ws[f'BQ{i}'] = val_bq_cc

        celda_br, celda_bt = ws[f'BR{i}'], ws[f'BT{i}']
        val_bs, val_bu = "", ""
        datos_maestro = maestro_dict.get(cedula_emp, (pd.NaT, pd.NaT))
        fecha_ing, fecha_ret = datos_maestro[0], datos_maestro[1]

        if pd.notna(fecha_ing):
            celda_br.value = fecha_ing.date()
            celda_br.number_format = 'DD/MM/YYYY'
            if fecha_ing > fecha_minima_valida and fecha_ori and pd.notna(fecha_ori) and fecha_ing > fecha_ori:
                val_bs = "Revisar"
        else:
            celda_br.value = ""

        if pd.notna(fecha_ret):
            celda_bt.value = fecha_ret.date()
            celda_bt.number_format = 'DD/MM/YYYY'
            if fecha_ret > fecha_minima_valida and fecha_ori and pd.notna(fecha_ori) and fecha_ret < fecha_ori:
                val_bu = "Retirado"
        else:
            celda_bt.value = ""

        ws[f'BS{i}'] = val_bs
        ws[f'BU{i}'] = val_bu

        val_bv_aus, val_bw_nova = val_bf, ""
        if val_bu == "Retirado":
            val_bv_aus = "Retiro"
        elif val_bs == "Revisar":
            val_bv_aus = "Ingreso"
        elif val_bo == "CCCO":
            val_bv_aus = "CCCO"
        elif f_val == "Descanso compensatorio":
            val_bv_aus = "C"
        elif cedula_emp in nova_dict and fecha_ori and pd.notna(fecha_ori):
            sub_nova = nova_dict[cedula_emp]
            match_nova = sub_nova[
                (sub_nova['Fecha_Inicio'] <= fecha_ori) &
                (sub_nova['Fecha_Fin'] >= fecha_ori)
            ]
            if not match_nova.empty:
                val_bv_aus = match_nova.iloc[0]['Concepto']
                val_bw_nova = match_nova.iloc[0]['Codigo_Novasoft']

        ws[f'BV{i}'] = val_bv_aus
        ws[f'BW{i}'] = val_bw_nova

        val_bx_sic = ""
        if cedula_emp in sic_dict and fecha_ori and pd.notna(fecha_ori):
            sub_sic = sic_dict[cedula_emp]
            match_sic = sub_sic[
                (sub_sic['Fecha_Inicio'] <= fecha_ori) &
                (sub_sic['Fecha_Fin'] >= fecha_ori)
            ]
            if not match_sic.empty:
                val_bx_sic = match_sic.iloc[0]['Proceso']

        ws[f'BX{i}'] = val_bx_sic

        val_by_consolidado = val_bv_aus
        if val_bv_aus == "Ausencia" and val_bx_sic != "":
            val_by_consolidado = val_bx_sic

        ws[f'BY{i}'] = val_by_consolidado

        # BZ: Compensado
        val_bz_comp = ""
        if cedula_emp and fecha_ori and pd.notna(fecha_ori):
            key_op = (cedula_emp, fecha_ori.date())
            val_bz_comp = operativa_dict.get(key_op, "")

        ws[f'BZ{i}'] = val_bz_comp

        # CA: Ausentismo Real
        val_ca_aus_real = val_by_consolidado
        if str(val_by_consolidado).strip().lower() in ["ausencia", "descanso", "p"] and str(val_bz_comp).strip().upper() == "C":
            val_ca_aus_real = "C"

        celda_ca = ws[f'CA{i}']
        celda_ca.value = val_ca_aus_real

        if cedula_emp and fecha_ori and pd.notna(fecha_ori):
            marcaciones_ht_dict[(cedula_emp, fecha_ori.date())] = f'=IFERROR(ROUND(BC{i}*24,1),"")'

        ca_val_clean = str(val_ca_aus_real).strip()
        if ca_val_clean.lower() == "ausencia":
            celda_ca.fill = fill_ausencia_rojo
            celda_ca.font = font_ausencia_blanco
            celda_ca.alignment = Alignment(horizontal="center", vertical="center")
            conteo_ausencias += 1

            nombre_emp = obtener_val_iloc(row, 1)
            f_str = fecha_ori.strftime('%d/%m/%Y') if fecha_ori and pd.notna(fecha_ori) else str(val_e)
            registros_ausencias.append({
                "Identificador": cedula_emp,
                "Nombres": nombre_emp,
                "Fecha Ori": f_str,
                "Dia": dia_nombre,
                "Ausentismo Real": "Ausencia"
            })
        elif ca_val_clean.upper() == "P":
            conteo_p += 1

        for col_letra in ['AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA']:
            ws[f'{col_letra}{i}'].border = thin_border

        pct = (idx + 1) / total_filas
        progress_bar.progress(pct)
        status_text.caption(f"⚡ Procesando fila {idx + 1} de {total_filas} ({int(pct*100)}%)")

    # ── CREACIÓN DE HOJA "Marcaciones Filtradas" (RANGO COMPLETO DE FECHAS POR TRABAJADOR) ──
    if fecha_ini_sup and fecha_fin_sup:
        nombre_hoja_filtrada = "Marcaciones Filtradas"
        if nombre_hoja_filtrada in wb.sheetnames:
            ws_m_filt = wb[nombre_hoja_filtrada]
            ws_m_filt.delete_rows(1, ws_m_filt.max_row + 1)
        else:
            ws_m_filt = wb.create_sheet(title=nombre_hoja_filtrada)
        
        ws_m_filt.views.sheetView[0].showGridLines = True

        max_col_m = ws.max_column
        for col_idx in range(1, max_col_m + 1):
            source_cell = ws.cell(row=1, column=col_idx)
            dest_cell = ws_m_filt.cell(row=1, column=col_idx, value=source_cell.value)
            
            # Copiar solo valores básicos de visualización para no arrastrar objetos StyleProxy de openpyxl
            dest_cell.fill = PatternFill(fill_type=None)
            dest_cell.font = Font(name="Calibri", size=10, bold=True)
            dest_cell.alignment = Alignment(horizontal="center", vertical="center")
            dest_cell.border = thin_border

        rango_fechas = pd.date_range(start=fecha_ini_sup, end=fecha_fin_sup).date
        rango_fechas_set = set(rango_fechas)

        marcaciones_por_emp = {}
        info_emp_dict = {}

        max_row_m = ws.max_row
        col_AY_idx = 51  # Columna AY (Fecha Ori)

        for r_idx in range(2, max_row_m + 1):
            ced_val = str(ws.cell(row=r_idx, column=3).value or '').strip().replace('.0', '')
            val_ay = ws.cell(row=r_idx, column=col_AY_idx).value

            f_dt = None
            if isinstance(val_ay, (datetime.date, datetime.datetime)):
                f_dt = val_ay if isinstance(val_ay, datetime.date) else val_ay.date()
            elif isinstance(val_ay, str) and len(val_ay) >= 10:
                try:
                    f_dt = pd.to_datetime(val_ay[-10:], dayfirst=True).date()
                except Exception:
                    pass

            if ced_val:
                if ced_val not in marcaciones_por_emp:
                    marcaciones_por_emp[ced_val] = {}
                    info_emp_dict[ced_val] = [ws.cell(row=r_idx, column=c).value for c in range(1, max_col_m + 1)]

                if f_dt and f_dt in rango_fechas_set:
                    marcaciones_por_emp[ced_val][f_dt] = r_idx

        idx_dest = 2
        for ced_val, fechas_dict in marcaciones_por_emp.items():
            for f_curr in rango_fechas:
                if f_curr in fechas_dict:
                    orig_r = fechas_dict[f_curr]
                    for col_idx in range(1, max_col_m + 1):
                        c_src = ws.cell(row=orig_r, column=col_idx)
                        c_dst = ws_m_filt.cell(row=idx_dest, column=col_idx, value=c_src.value)
                        c_dst.number_format = c_src.number_format
                        c_dst.border = thin_border
                else:
                    base_data = info_emp_dict[ced_val].copy()
                    for col_idx in range(1, max_col_m + 1):
                        val_dummy = base_data[col_idx - 1] if col_idx <= len(base_data) else ""
                        c_dst = ws_m_filt.cell(row=idx_dest, column=col_idx)
                        c_dst.border = thin_border

                        if col_idx == 51:  # Columna AY (Fecha Ori)
                            c_dst.value = f_curr
                            c_dst.number_format = 'DD/MM/YYYY'
                        elif col_idx == 52:  # Columna AZ (Día)
                            dia_n = dias_semana_es[f_curr.weekday()]
                            if f_curr in set_festivos:
                                dia_n += " Festivo"
                            c_dst.value = dia_n
                        else:
                            c_dst.value = val_dummy

                idx_dest += 1

    # Creación de la hoja "Ausencias"
    if "Ausencias" in wb.sheetnames:
        ws_aus = wb["Ausencias"]
        ws_aus.delete_rows(1, ws_aus.max_row + 1)
    else:
        ws_aus = wb.create_sheet(title="Ausencias")
    
    ws_aus.views.sheetView[0].showGridLines = True

    bg_azul_header = PatternFill(start_color="1B5E82", end_color="1B5E82", fill_type="solid")
    bg_verde_header = PatternFill(start_color="1E4620", end_color="1E4620", fill_type="solid")
    bg_azul_dark = PatternFill(start_color="00529B", end_color="00529B", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    ws_aus["A1"] = "Identificador"
    ws_aus["B1"] = "Nombres"
    ws_aus["C1"] = "Fecha Ori"
    ws_aus["D1"] = "Dia"
    ws_aus["E1"] = "Ausentismo Real"

    for col in ["A1", "B1", "C1", "D1"]:
        c = ws_aus[col]
        c.fill = bg_azul_header
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    c_e1 = ws_aus["E1"]
    c_e1.fill = bg_verde_header
    c_e1.font = font_header
    c_e1.alignment = Alignment(horizontal="center", vertical="center")
    c_e1.border = thin_border

    df_aus_det = pd.DataFrame(registros_ausencias)
    for row_idx, r in enumerate(registros_ausencias, start=2):
        ws_aus[f"A{row_idx}"] = r["Identificador"]
        ws_aus[f"B{row_idx}"] = r["Nombres"]
        ws_aus[f"C{row_idx}"] = r["Fecha Ori"]
        ws_aus[f"D{row_idx}"] = r["Dia"]
        
        c_aus_val = ws_aus[f"E{row_idx}"]
        c_aus_val.value = r["Ausentismo Real"]
        c_aus_val.fill = fill_ausencia_rojo
        c_aus_val.font = font_ausencia_blanco
        c_aus_val.alignment = Alignment(horizontal="center", vertical="center")

        for col_l in ["A", "B", "C", "D", "E"]:
            ws_aus[f"{col_l}{row_idx}"].border = thin_border

    ws_aus["H1"] = "Identificador"
    ws_aus["I1"] = "Cantidad"

    for col in ["H1", "I1"]:
        c = ws_aus[col]
        c.fill = bg_azul_dark
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    if not df_aus_det.empty:
        df_resumen = df_aus_det.groupby("Identificador").size().reset_index(name="Cantidad")
        df_resumen = df_resumen.sort_values(by="Cantidad", ascending=False)

        for r_idx, r in enumerate(df_resumen.itertuples(), start=2):
            c_h = ws_aus[f"H{r_idx}"]
            c_i = ws_aus[f"I{r_idx}"]
            
            c_h.value = r.Identificador
            c_i.value = r.Cantidad
            
            c_h.alignment = Alignment(horizontal="center", vertical="center")
            c_i.alignment = Alignment(horizontal="right", vertical="center")
            c_h.border = thin_border
            c_i.border = thin_border

    # Creación y cruce de la hoja "Supernumerario"
    if not df_super_filtrado.empty:
        if "Supernumerario" in wb.sheetnames:
            ws_sup = wb["Supernumerario"]
            ws_sup.delete_rows(1, ws_sup.max_row + 1)
        else:
            ws_sup = wb.create_sheet(title="Supernumerario")
        
        ws_sup.views.sheetView[0].showGridLines = True
        
        for r_idx, r in enumerate(dataframe_to_rows(df_super_filtrado, index=False, header=True), start=1):
            ws_sup.append(r)
            for c_idx in range(1, len(r) + 1):
                cell = ws_sup.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                if r_idx == 1:
                    cell.fill = PatternFill(start_color="00529B", end_color="00529B", fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_sup.cell(row=1, column=23, value="HT")
        ws_sup.cell(row=1, column=23).fill = PatternFill(start_color="00529B", end_color="00529B", fill_type="solid")
        ws_sup.cell(row=1, column=23).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ws_sup.cell(row=1, column=23).alignment = Alignment(horizontal="center", vertical="center")

        max_r_sup = ws_sup.max_row
        for r_sup in range(2, max_r_sup + 1):
            cell_f = ws_sup.cell(row=r_sup, column=2)
            f_val_sup = cell_f.value
            dt_sup = None
            
            if pd.notna(f_val_sup):
                try:
                    dt_sup = pd.to_datetime(f_val_sup, dayfirst=True, errors='coerce')
                    if pd.notna(dt_sup):
                        cell_f.value = dt_sup.date()
                        cell_f.number_format = 'DD/MM/YYYY'
                except Exception:
                    pass

            cell_ced = ws_sup.cell(row=r_sup, column=3)
            ced_sup = str(cell_ced.value).strip().replace('.0', '') if cell_ced.value else ""

            cell_ht = ws_sup.cell(row=r_sup, column=23)
            cell_ht.border = thin_border
            cell_ht.alignment = Alignment(horizontal="center", vertical="center")

            if ced_sup and dt_sup and pd.notna(dt_sup):
                key_sup = (ced_sup, dt_sup.date())
                ht_encontrado = marcaciones_ht_dict.get(key_sup, "")
                cell_ht.value = ht_encontrado
            else:
                cell_ht.value = ""

    status_text.empty()
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, conteo_ausencias, conteo_p, total_filas


# ─── INTERFAZ DE USUARIO ───────────────────────────────────────────────────

# Configuración del Panel Lateral
st.sidebar.markdown("## ⚙️ Parámetros")
contrato_principal = st.sidebar.text_input("Contrato / CC Principal", value="FUNDACION HOSPITAL DE LA MISERICORDIA")

st.sidebar.markdown("---")
st.sidebar.markdown("### 📅 Filtro Rango de Fechas")
fecha_ini_sup = st.sidebar.date_input("Fecha Inicial", value=datetime.date(2026, 2, 1))
fecha_fin_sup = st.sidebar.date_input("Fecha Final", value=datetime.date(2026, 2, 28))

st.sidebar.markdown("---")
st.sidebar.markdown("""
<div style="background-color: #f0f7ff; padding: 12px; border-radius: 8px; border-left: 4px solid #00529B;">
    <small style="color: #00529B; font-weight: 600;">💡 Instrucciones</small><br>
    <small style="color: #475569;">1. Despliega 'Bases de datos' y carga los archivos.<br>2. Ajusta las fechas del periodo.<br>3. Ejecuta la auditoría.</small>
</div>
""", unsafe_allow_html=True)

# ── 1. ACORDEÓN PRINCIPAL: "Bases de datos" ──
with st.expander("📁 Bases de datos", expanded=True):
    col1, col2 = st.columns(2, gap="large")

    with col1:
        file_entrada = st.file_uploader("1. Marcaciones GeoVictoria (.xlsx)", type=["xlsx"])
        file_operativa = st.file_uploader("2. Base Operativa (.xlsx)", type=["xlsx"])
        file_novasoft = st.file_uploader("3. BBDD Novasoft (.xlsx)", type=["xlsx"])
        file_supernumerario = st.file_uploader("7. BD Supernumerario (.xlsx)", type=["xlsx"])

        status_e = '<span class="file-status-ok">✔ Principal Cargado</span>' if file_entrada else '<span class="file-status-pending">Pendiente Marcaciones</span>'

        st.markdown(f"""
            <div class="card-container">
                <div class="section-header">
                    <span>📌 Archivos Principales & Operativos</span>
                    {status_e}
                </div>
            </div>
        """, unsafe_allow_html=True)

    with col2:
        file_sic = st.file_uploader("4. Informe SIC (.xlsx)", type=["xlsx"])
        file_maestro = st.file_uploader("5. Base Maestro (.xlsx)", type=["xlsx"])
        file_historial = st.file_uploader("6. Historial Laboral (.xlsx)", type=["xlsx"])

        count_comp = sum(1 for x in [file_sic, file_maestro, file_historial, file_supernumerario] if x is not None)
        status_c = f'<span class="file-status-ok">✔ {count_comp}/4 Cargados</span>' if count_comp > 0 else '<span class="file-status-pending">Opcionales</span>'

        st.markdown(f"""
            <div class="card-container">
                <div class="section-header">
                    <span>📊 Bases Complementarias</span>
                    {status_c}
                </div>
            </div>
        """, unsafe_allow_html=True)

# ── 2. ACORDEÓN SECUNDARIO: Configuración Avanzada ──
with st.expander("🛠️ Configuración Avanzada de Pestañas (Opcional)"):
    st.caption("Solo modifica estos campos si los libros de Excel tienen nombres de hoja diferentes a los estándar.")
    c_a, c_b = st.columns(2)
    with c_a:
        hoja_entrada = st.text_input("1. Hoja Marcaciones", value="Marcaciones")
        hoja_festivos = st.text_input("Hoja Festivos", value="Festivos")
        hoja_operativa = st.text_input("2. Hoja Operativa", value="CONSOLIDADO")
        hoja_novasoft = st.text_input("3. Hoja Novasoft", value="BBDD_Novasof")
    with c_b:
        hoja_sic = st.text_input("4. Hoja SIC", value="Datos")
        hoja_maestro = st.text_input("5. Hoja Maestro", value="NOM1911")
        hoja_historial = st.text_input("6. Hoja Historial", value="Hoja 1")
        hoja_supernumerario = st.text_input("7. Hoja Supernumerario", value="Base")

st.markdown("<br>", unsafe_allow_html=True)

# Botón de Procesamiento
if st.button("⚡ Ejecutar Auditoría TS y Procesar Marcaciones", type="primary"):
    if not file_entrada:
        st.error("⚠️ Es obligatorio cargar el archivo principal de Marcaciones (GeoVictoria).")
    elif not contrato_principal:
        st.error("⚠️ Por favor, ingresa el valor del Contrato Principal en el panel izquierdo.")
    else:
        try:
            with st.spinner("Procesando marcaciones, generando hoja 'Marcaciones Filtradas' y aplicando estilos..."):
                excel_salida, kpi_ausencias, kpi_p, total_filas = procesar_plantilla_geovictoria(
                    file_entrada, hoja_entrada, sheet_festivos=hoja_festivos,
                    file_operativa=file_operativa, sheet_operativa=hoja_operativa,
                    file_novasoft=file_novasoft, sheet_novasoft=hoja_novasoft,
                    file_sic=file_sic, sheet_sic=hoja_sic,
                    file_maestro=file_maestro, sheet_maestro=hoja_maestro,
                    file_historial=file_historial, sheet_historial=hoja_historial,
                    file_supernumerario=file_supernumerario, sheet_supernumerario=hoja_supernumerario,
                    contrato_principal=contrato_principal,
                    fecha_ini_sup=fecha_ini_sup, fecha_fin_sup=fecha_fin_sup
                )

            st.session_state["procesado_exitoso"] = True
            st.session_state["excel_salida"] = excel_salida
            st.session_state["kpi_ausencias"] = kpi_ausencias
            st.session_state["kpi_p"] = kpi_p
            st.session_state["total_filas"] = total_filas

        except Exception as e:
            st.error(f"❌ Ocurrió un error durante el procesamiento: {str(e)}")

# ── RENDERIZADO PERSISTENTE DE RESULTADOS Y KPIS ──
if st.session_state.get("procesado_exitoso", False):
    st.success("✨ ¡Auditoría finalizada con éxito! Hoja 'Marcaciones Filtradas' creada asegurando la secuencia completa de fechas por trabajador.")
    
    st.download_button(
        label="📥 Descargar Resultado Calculado (Excel)",
        data=st.session_state["excel_salida"],
        file_name="Calculado_GeoVictoria_Casalimpia.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.markdown("<br><h3 style='color: #00529B; font-weight: 700;'>📊 Resumen Ejecutivo de Auditoría</h3>", unsafe_allow_html=True)
    kpi_col1, kpi_col2, kpi_col3 = st.columns(3)

    with kpi_col1:
        st.markdown(f"""
            <div class="kpi-card kpi-card-danger">
                <div class="kpi-title kpi-title-danger">🚨 Ausencias Reales (CA)</div>
                <div class="kpi-value kpi-value-danger">{st.session_state["kpi_ausencias"]:,}</div>
                <div class="kpi-subtitle">Registros clasificados como Ausencia en Columna CA</div>
            </div>
        """, unsafe_allow_html=True)

    with kpi_col2:
        st.markdown(f"""
            <div class="kpi-card kpi-card-warning">
                <div class="kpi-title kpi-title-warning">⚠️ Marcaciones Erróneas (P)</div>
                <div class="kpi-value kpi-value-warning">{st.session_state["kpi_p"]:,}</div>
                <div class="kpi-subtitle">Registros marcados con P en Columna CA</div>
            </div>
        """, unsafe_allow_html=True)

    with kpi_col3:
        st.markdown(f"""
            <div class="kpi-card kpi-card-info">
                <div class="kpi-title kpi-title-info">📋 Total Registros Procesados</div>
                <div class="kpi-value kpi-value-info">{st.session_state["total_filas"]:,}</div>
                <div class="kpi-subtitle">Filas evaluadas en el periodo</div>
            </div>
        """, unsafe_allow_html=True)
